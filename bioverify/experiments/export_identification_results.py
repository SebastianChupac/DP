# export_identification_results.py - Export identification experiment results to CSV, JSON, and XLSX.
# Author: Sebastian Chupac (xchupa03)
# Date: 19.05.2026



"""
Export identification experiment results from bioverify/results/identification to tabular formats.
Generates CSV, JSON, and XLSX files with full numeric precision.

Usage:
    python export_identification_results.py
    python export_identification_results.py --output bioverify/results/id_results
"""

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font


# Type hint for extraction results
ProbeGalleryCounts = Tuple[Optional[int], Optional[int]]


class IdentificationExporter:
    """Export identification results to multiple formats with consistent schema."""

    FLOAT_COLUMNS = {
        'rank_1_accuracy', 'rank_5_accuracy', 'rank_10_accuracy', 'rank_15_accuracy',
        'mean_average_precision', 'mean_rank', 'median_rank', 'std_rank'
    }
    INTEGER_COLUMNS = {'num_probes', 'num_valid_ranks', 'shortlist_k', 
                       'probes_per_identity', 'gallery_samples_per_identity'}
    
    COLUMN_ORDER = [
        'experiment_folder', 'experiment_title', 'base_matcher', 'matcher',
        'modality', 'dataset', 'dataset_file', 'config_file', 'summary_file',
        'ranking_strategy', 'shortlist_k', 'shortlist_matcher',
        'probes_per_identity', 'gallery_samples_per_identity',
        'num_probes', 'num_valid_ranks', 'rank_1_accuracy', 'rank_5_accuracy',
        'rank_10_accuracy', 'rank_15_accuracy', 'mean_average_precision',
        'mean_rank', 'median_rank', 'std_rank'
    ]

    def __init__(self, results_root: Path):
        self.results_root = results_root / 'identification'
        
    def collect_identification_rows(self) -> List[Dict[str, Any]]:
        """Collect all identification result rows from experiment folders."""
        rows = []
        if not self.results_root.exists():
            raise FileNotFoundError(f"Results directory not found: {self.results_root}")
        
        experiment_folders = sorted([p for p in self.results_root.iterdir() if p.is_dir()])
        
        for exp_folder in experiment_folders:
            config_file = exp_folder / 'config.yaml'
            summary_file = exp_folder / 'summary.json'
            
            if not config_file.exists() or not summary_file.exists():
                print(f"Skipping {exp_folder.name}: missing config.yaml or summary.json")
                continue
            
            try:
                config = yaml.safe_load(config_file.read_text(encoding='utf-8'))
                summary = json.loads(summary_file.read_text(encoding='utf-8'))
            except Exception as e:
                print(f"Error reading {exp_folder.name}: {e}")
                continue
            
            # Extract experiment metadata
            exp_title = config.get('experiment', {}).get('name', '')
            ranking_strategy = config.get('ranking_strategy', '')
            shortlist_k = config.get('shortlist_k', '')
            
            # Handle shortlist_matcher which can be str, dict, or None
            shortlist_matcher_raw = config.get('shortlist_matcher')
            if isinstance(shortlist_matcher_raw, dict):
                shortlist_matcher = shortlist_matcher_raw.get('name', '')
            elif isinstance(shortlist_matcher_raw, str):
                shortlist_matcher = shortlist_matcher_raw
            else:
                shortlist_matcher = ''
            
            dataset_file = config.get('identification_dataset', '')
            probes_per_identity, gallery_samples_per_identity = self._extract_probe_gallery_counts(dataset_file)
            
            # For each matcher in summary, create a row
            for matcher_name, metrics in summary.items():
                # Skip non-matcher entries (shouldn't be any, but be safe)
                if not isinstance(metrics, dict) or 'rank_1_accuracy' not in metrics:
                    continue
                
                base_matcher = self._infer_base_matcher(matcher_name)
                dataset = self._infer_dataset_name(dataset_file, config)
                modality = self._infer_modality(exp_folder.name, dataset_file, config)
                
                # Extract rank accuracy for specific k values
                rank_k_accuracy = metrics.get('rank_k_accuracy', {})
                
                row = {
                    'experiment_folder': exp_folder.name,
                    'experiment_title': exp_title,
                    'base_matcher': base_matcher,
                    'matcher': matcher_name,
                    'modality': modality,
                    'dataset': dataset,
                    'dataset_file': dataset_file,
                    'config_file': str(config_file.relative_to(self.results_root.parent.parent)),
                    'summary_file': str(summary_file.relative_to(self.results_root.parent.parent)),
                    'ranking_strategy': ranking_strategy,
                    'shortlist_k': shortlist_k,
                    'shortlist_matcher': shortlist_matcher,
                    'probes_per_identity': probes_per_identity,
                    'gallery_samples_per_identity': gallery_samples_per_identity,
                    'num_probes': metrics.get('num_probes'),
                    'num_valid_ranks': metrics.get('num_valid_ranks'),
                    'rank_1_accuracy': rank_k_accuracy.get('1'),
                    'rank_5_accuracy': rank_k_accuracy.get('5'),
                    'rank_10_accuracy': rank_k_accuracy.get('10'),
                    'rank_15_accuracy': rank_k_accuracy.get('15'),
                    'mean_average_precision': metrics.get('mean_average_precision'),
                    'mean_rank': metrics.get('mean_rank'),
                    'median_rank': metrics.get('median_rank'),
                    'std_rank': metrics.get('std_rank'),
                }
                rows.append(row)
        
        return rows

    def _extract_probe_gallery_counts(self, dataset_file: str) -> Tuple[Optional[int], Optional[int]]:
        """Extract probes per identity and gallery samples per identity from dataset filename.
        
        Format: identification_{modality}_{identities}_{probes}_{gallery}_{id_range}.csv
        Returns: (probes_per_identity, gallery_samples_per_identity) or (None, None) if not found
        """
        filename = Path(dataset_file).stem
        match = re.search(r'_(\d+)_(\d+)_(\d+)_', filename)
        if match:
            # identities = match.group(1)  # not needed
            probes = int(match.group(2))
            gallery = int(match.group(3))
            return probes, gallery
        return None, None

    def _infer_base_matcher(self, matcher_name: str) -> str:
        """Extract base matcher class from matcher name."""
        matcher_lower = matcher_name.lower()
        for base in ['aspanformer', 'deepdetect', 'sgmnet', 'superglue', 'loftr', 'sift', 'orb']:
            if base in matcher_lower:
                return base
        return matcher_name.lower()

    def _infer_dataset_name(self, dataset_file: str, config: Dict) -> str:
        """Extract dataset name from dataset file or config."""
        modality_tokens = {'face', 'iris', 'hand', 'finger', 'fingervein'}
        stopwords = {
            'identification', 'pairs', 'test', 'train', 'validation', 'masked', 'mask',
            'clahe', 'new', 'score', 'indoor', 'outdoor', 'angle', 'lighting',
            'expression', 'change', 'csv', 'left', 'right', 'middle', 'index', 'ring',
            'pinky', 'dorsal', 'palmar', 'thumb', 'single', 'multiple', 'fixed',
            'crop', 'cropped'
        }

        def _infer_from_text(value: str) -> str:
            tokens = [token for token in re.split(r'[_-]+', value.lower()) if token]
            if tokens and tokens[0] == 'identification':
                tokens = tokens[1:]
            if tokens and tokens[0] in modality_tokens:
                tokens = tokens[1:]

            for token in tokens:
                if token in stopwords:
                    continue
                if token.isdigit():
                    break
                return token

            return ''

        # Try extracting from dataset file path
        if dataset_file:
            dataset_name = _infer_from_text(Path(dataset_file).stem)
            if dataset_name:
                return dataset_name

        # Try extracting from experiment name
        exp_name = config.get('experiment', {}).get('name', '')
        if isinstance(exp_name, str) and exp_name.strip():
            dataset_name = _infer_from_text(exp_name)
            if dataset_name:
                return dataset_name

        return ''

    def _infer_modality(self, folder_name: str, dataset_file: str, config: Dict) -> str:
        """Infer modality from dataset file, config, or folder name."""
        filter_modality = config.get('filter_modality')
        if filter_modality:
            return filter_modality
        
        # Check dataset file
        if dataset_file:
            dataset_lower = dataset_file.lower()
            if 'face' in dataset_lower:
                return 'face'
            if 'iris' in dataset_lower:
                return 'iris'
            if 'hand' in dataset_lower:
                return 'hand'
            if 'finger' in dataset_lower or 'fingervein' in dataset_lower:
                return 'fingervein'
        
        # Check experiment name
        exp_name = config.get('experiment', {}).get('name', '').lower()
        if 'face' in exp_name:
            return 'face'
        if 'iris' in exp_name:
            return 'iris'
        if 'hand' in exp_name:
            return 'hand'
        if 'finger' in exp_name or 'fingervein' in exp_name:
            return 'fingervein'
        
        # Check folder name
        folder_lower = folder_name.lower()
        if 'face' in folder_lower:
            return 'face'
        if 'iris' in folder_lower:
            return 'iris'
        if 'hand' in folder_lower:
            return 'hand'
        if 'finger' in folder_lower or 'fingervein' in folder_lower:
            return 'fingervein'
        
        return ''

    def write_csv(self, rows: List[Dict], output_path: Path) -> None:
        """Write results to CSV with full numeric precision."""
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=self.COLUMN_ORDER)
            writer.writeheader()
            writer.writerows(rows)

    def write_json(self, rows: List[Dict], output_path: Path) -> None:
        """Write results to JSON preserving full numeric precision."""
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(rows, f, indent=2)

    def write_xlsx(self, rows: List[Dict], output_path: Path) -> None:
        """Write results to XLSX with explicit formatting."""
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Identification Results"
        
        # Header
        ws.append(self.COLUMN_ORDER)
        header_row = ws[1]
        for cell in header_row:
            cell.font = Font(bold=True)
        
        # Data rows
        for row_dict in rows:
            row_data = [row_dict.get(col) for col in self.COLUMN_ORDER]
            ws.append(row_data)
            
            # Apply formatting
            for idx, col_name in enumerate(self.COLUMN_ORDER, 1):
                cell = ws.cell(row=ws.max_row, column=idx)
                if col_name in self.FLOAT_COLUMNS:
                    cell.number_format = '0.###############'
                elif col_name in self.INTEGER_COLUMNS:
                    cell.number_format = '0'
        
        # Adjust column widths
        for idx, col_name in enumerate(self.COLUMN_ORDER, 1):
            ws.column_dimensions[chr(64 + idx if idx <= 26 else 64 + idx)].width = 20
        
        wb.save(output_path)

    def export_experiment(self, exp_folder: Path) -> Tuple[int, Path, Path, Path]:
        """Export a single experiment folder and write results into that folder.
        Returns (row_count, csv_path, json_path, xlsx_path).
        """
        exp_folder = Path(exp_folder)
        if not exp_folder.exists() or not exp_folder.is_dir():
            raise FileNotFoundError(f"Experiment folder not found: {exp_folder}")

        config_file = exp_folder / 'config.yaml'
        summary_file = exp_folder / 'summary.json'
        if not config_file.exists() or not summary_file.exists():
            raise FileNotFoundError(f"Missing config.yaml or summary.json in {exp_folder}")

        try:
            config = yaml.safe_load(config_file.read_text(encoding='utf-8'))
            summary = json.loads(summary_file.read_text(encoding='utf-8'))
        except Exception as e:
            raise RuntimeError(f"Failed to read experiment files: {e}")

        exp_title = config.get('experiment', {}).get('name', '')
        ranking_strategy = config.get('ranking_strategy', '')
        shortlist_k = config.get('shortlist_k', '')
        shortlist_matcher_raw = config.get('shortlist_matcher')
        if isinstance(shortlist_matcher_raw, dict):
            shortlist_matcher = shortlist_matcher_raw.get('name', '')
        elif isinstance(shortlist_matcher_raw, str):
            shortlist_matcher = shortlist_matcher_raw
        else:
            shortlist_matcher = ''

        dataset_file = config.get('identification_dataset', '')
        probes_per_identity, gallery_samples_per_identity = self._extract_probe_gallery_counts(dataset_file)

        rows: List[Dict[str, Any]] = []
        for matcher_name, metrics in summary.items():
            if not isinstance(metrics, dict) or 'rank_1_accuracy' not in metrics:
                continue
            base_matcher = self._infer_base_matcher(matcher_name)
            dataset = self._infer_dataset_name(dataset_file, config)
            modality = self._infer_modality(exp_folder.name, dataset_file, config)
            rank_k_accuracy = metrics.get('rank_k_accuracy', {})

            row = {
                'experiment_folder': exp_folder.name,
                'experiment_title': exp_title,
                'base_matcher': base_matcher,
                'matcher': matcher_name,
                'modality': modality,
                'dataset': dataset,
                'dataset_file': dataset_file,
                'config_file': str(config_file.relative_to(self.results_root.parent.parent)),
                'summary_file': str(summary_file.relative_to(self.results_root.parent.parent)),
                'ranking_strategy': ranking_strategy,
                'shortlist_k': shortlist_k,
                'shortlist_matcher': shortlist_matcher,
                'probes_per_identity': probes_per_identity,
                'gallery_samples_per_identity': gallery_samples_per_identity,
                'num_probes': metrics.get('num_probes'),
                'num_valid_ranks': metrics.get('num_valid_ranks'),
                'rank_1_accuracy': rank_k_accuracy.get('1'),
                'rank_5_accuracy': rank_k_accuracy.get('5'),
                'rank_10_accuracy': rank_k_accuracy.get('10'),
                'rank_15_accuracy': rank_k_accuracy.get('15'),
                'mean_average_precision': metrics.get('mean_average_precision'),
                'mean_rank': metrics.get('mean_rank'),
                'median_rank': metrics.get('median_rank'),
                'std_rank': metrics.get('std_rank'),
            }
            rows.append(row)

        base_name = f"{exp_folder.name}_identification_results"
        csv_path = exp_folder / f"{base_name}.csv"
        json_path = exp_folder / f"{base_name}.json"
        xlsx_path = exp_folder / f"{base_name}.xlsx"

        self.write_csv(rows, csv_path)
        self.write_json(rows, json_path)
        self.write_xlsx(rows, xlsx_path)

        return len(rows), csv_path, json_path, xlsx_path

    def export_all(self, output_dir: Optional[Path] = None) -> Tuple[int, Path, Path, Path]:
        """Export to all formats. Returns (row_count, csv_path, json_path, xlsx_path)."""
        if output_dir is None:
            output_dir = self.results_root.parent
        else:
            output_dir = Path(output_dir)
        
        output_dir.mkdir(parents=True, exist_ok=True)
        
        print("Collecting identification results...")
        rows = self.collect_identification_rows()
        print(f"Found {len(rows)} rows across all matchers")
        
        csv_path = output_dir / 'identification_results.csv'
        json_path = output_dir / 'identification_results.json'
        xlsx_path = output_dir / 'identification_results.xlsx'
        
        print(f"Writing CSV: {csv_path}")
        self.write_csv(rows, csv_path)
        
        print(f"Writing JSON: {json_path}")
        self.write_json(rows, json_path)
        
        print(f"Writing XLSX: {xlsx_path}")
        self.write_xlsx(rows, xlsx_path)
        
        return len(rows), csv_path, json_path, xlsx_path


def main():
    parser = argparse.ArgumentParser(
        description='Export identification experiment results to CSV/JSON/XLSX'
    )
    parser.add_argument(
        '--output',
        type=str,
        default=None,
        help='Output directory (default: bioverify/results)'
    )
    parser.add_argument(
        '--experiment',
        type=str,
        default=None,
        help='Path to a single experiment folder to export results for (writes files inside that folder)'
    )
    
    args = parser.parse_args()
    
    # Determine results root
    current_file = Path(__file__)
    src_root = current_file.parent.parent.parent
    results_root = src_root / 'bioverify' / 'results'
    
    exporter = IdentificationExporter(results_root)
    output_dir = Path(args.output) if args.output else results_root
    
    if args.experiment:
        exp_path = Path(args.experiment)
        row_count, csv_path, json_path, xlsx_path = exporter.export_experiment(exp_path)
    else:
        row_count, csv_path, json_path, xlsx_path = exporter.export_all(output_dir)
    
    print(f"\nExport complete!")
    print(f"  Total rows: {row_count}")
    print(f"  CSV:  {csv_path}")
    print(f"  JSON: {json_path}")
    print(f"  XLSX: {xlsx_path}")


if __name__ == '__main__':
    main()

