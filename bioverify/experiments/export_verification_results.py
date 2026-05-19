# export_verification_results.py - Export verification experiment results to CSV, JSON, and XLSX.
# Author: Sebastian Chupac (xchupa03)
# Date: 19.05.2026

"""Export verification experiment results into tabular files.

The exporter scans ``bioverify/results`` for verification experiment folders,
skips the identification subtree, and writes one row per matcher per experiment.
Core metrics come from ``summary.json`` and ROC AUC is pulled from the matching
``evaluation/*_threshold_analysis.json`` file when present.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


OUTPUT_COLUMNS = [
    "experiment_folder",
    "experiment_title",
    "base_matcher",
    "matcher",
    "modality",
    "dataset",
    "dataset_file",
    "config_file",
    "summary_file",
    "threshold_analysis_file",
    "eer",
    "accuracy",
    "precision",
    "recall",
    "roc_auc",
    "tar",
    "far",
    "frr",
    "trr",
    "num_pairs",
    "num_genuine",
    "num_impostor",
    "avg_inlier_ratio",
]

FLOAT_COLUMNS = {
    "accuracy",
    "precision",
    "recall",
    "roc_auc",
    "eer",
    "tar",
    "far",
    "frr",
    "trr",
    "avg_inlier_ratio",
}

INTEGER_COLUMNS = {
    "num_pairs",
    "num_genuine",
    "num_impostor",
}

BASE_MATCHER_PATTERNS = [
    ("aspanformer", "aspanformer"),
    ("deepdetect", "deepdetect"),
    ("sgmnet", "sgmnet"),
    ("superglue", "superglue"),
    ("loftr", "loftr"),
    ("sift", "sift"),
    ("orb", "orb"),
]

DATASET_STOPWORDS = {
    "pairs",
    "pair",
    "test",
    "train",
    "validation",
    "val",
    "all",
    "matchers",
    "matcher",
    "masked",
    "mask",
    "nomask",
    "without",
    "with",
    "crop",
    "cropped",
    "clahe",
    "new",
    "score",
    "override",
    "old",
    "root",
    "indoor",
    "outdoor",
    "angle",
    "expression",
    "lighting",
    "change",
    "version",
    "nc",
    "general",
    "side",
    "palmar",
    "dorsal",
    "fingervein",
    "finger",
    "roi",
    "rois",
    "enhancement",
    "enhacement",
    "matched",
}

MODALITY_PATTERNS = [
    ("face", "face"),
    ("iris", "iris"),
    ("hands", "hand"),
    ("hand", "hand"),
    ("fingervein", "fingervein"),
    ("finger", "finger"),
]


def _project_root() -> Path:
    return Path(__file__).resolve().parents[2]


def _load_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def _load_yaml(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        data = yaml.safe_load(handle)
    return data or {}


def _normalize_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def _split_tokens(value: str) -> List[str]:
    return [token for token in re.split(r"[^a-z0-9]+", value.lower()) if token]


def _strip_dataset_tokens(tokens: List[str]) -> List[str]:
    stripped = list(tokens)
    while stripped and (stripped[0] in DATASET_STOPWORDS or stripped[0].isdigit() or re.fullmatch(r"\d+k", stripped[0])):
        stripped.pop(0)

    while stripped and (stripped[-1] in DATASET_STOPWORDS or stripped[-1].isdigit()):
        stripped.pop()

    return stripped


def _infer_base_matcher(matcher_name: str) -> str:
    normalized = matcher_name.lower()
    for token, base_name in BASE_MATCHER_PATTERNS:
        if token in normalized:
            return base_name
    return _normalize_name(matcher_name)


def _infer_dataset_name_from_text(value: str, modality: str = "") -> str:
    tokens = _strip_dataset_tokens(_split_tokens(value))
    if not tokens:
        return ""

    if modality:
        modality_tokens = _split_tokens(modality)
        for index, token in enumerate(tokens):
            if token in modality_tokens and index + 1 < len(tokens):
                candidate = tokens[index + 1]
                if candidate not in DATASET_STOPWORDS:
                    return candidate

    if len(tokens) == 1:
        return tokens[0]

    for token in tokens:
        if token not in DATASET_STOPWORDS and not token.isdigit():
            return token

    return tokens[0]


def _infer_dataset_name(folder_name: str, config: Dict[str, Any], modality: str) -> str:
    dataset = config.get("dataset")
    if isinstance(dataset, str) and dataset.strip():
        dataset_name = _infer_dataset_name_from_text(Path(dataset).stem, modality)
        if dataset_name:
            return dataset_name

    experiment = config.get("experiment", {})
    if isinstance(experiment, dict):
        experiment_name = experiment.get("name")
        if isinstance(experiment_name, str) and experiment_name.strip():
            dataset_name = _infer_dataset_name_from_text(experiment_name, modality)
            if dataset_name:
                return dataset_name

    return _infer_dataset_name_from_text(folder_name, modality)


def _infer_modality(folder_name: str, config: Dict[str, Any]) -> str:
    candidates: List[str] = []

    filter_modality = config.get("filter_modality")
    if isinstance(filter_modality, str) and filter_modality.strip():
        candidates.append(filter_modality)

    dataset = config.get("dataset")
    if isinstance(dataset, str) and dataset.strip():
        candidates.append(Path(dataset).stem)

    experiment = config.get("experiment", {})
    if isinstance(experiment, dict):
        experiment_name = experiment.get("name")
        if isinstance(experiment_name, str) and experiment_name.strip():
            candidates.append(experiment_name)

    candidates.append(folder_name)

    for candidate in candidates:
        search_space = candidate.lower()
        for token, modality in MODALITY_PATTERNS:
            if token in search_space:
                return modality

    return ""


def _find_threshold_analysis_file(evaluation_dir: Path, matcher_name: str) -> Optional[Path]:
    if not evaluation_dir.exists():
        return None

    target = _normalize_name(matcher_name)
    candidates = list(evaluation_dir.glob("*_threshold_analysis.json"))
    for candidate in candidates:
        if _normalize_name(candidate.stem.replace("_threshold_analysis", "")) == target:
            return candidate

    for candidate in candidates:
        if target and target in _normalize_name(candidate.stem):
            return candidate

    return None


def _extract_auc(threshold_analysis_path: Optional[Path]) -> Optional[float]:
    if threshold_analysis_path is None or not threshold_analysis_path.exists():
        return None

    try:
        data = _load_json(threshold_analysis_path)
    except json.JSONDecodeError:
        return None

    roc = data.get("roc")
    if isinstance(roc, dict) and roc.get("auc") is not None:
        try:
            return float(roc["auc"])
        except (TypeError, ValueError):
            return None
    return None


def _extract_eer(evaluation_dir: Path, matcher_name: str) -> Optional[float]:
    """Load evaluation/eer_comparison.json and return EER for matcher if available."""
    eer_path = evaluation_dir / "eer_comparison.json"
    if not eer_path.exists():
        return None

    try:
        data = _load_json(eer_path)
    except json.JSONDecodeError:
        return None

    if not isinstance(data, dict):
        return None

    # Prefer base matcher key
    base = _infer_base_matcher(matcher_name)
    # direct match
    if base in data and isinstance(data[base], dict) and data[base].get("eer") is not None:
        try:
            return float(data[base]["eer"])
        except (TypeError, ValueError):
            return None

    # try case-insensitive match and normalized name match
    target_norm = _normalize_name(matcher_name)
    for key, val in data.items():
        if not isinstance(val, dict):
            continue
        key_norm = _normalize_name(key)
        if key_norm == target_norm or key_norm == base or base in key_norm or key_norm in base:
            if val.get("eer") is not None:
                try:
                    return float(val["eer"])
                except (TypeError, ValueError):
                    return None

    return None


def _format_relative_path(path: Path, project_root: Path) -> str:
    try:
        return path.relative_to(project_root).as_posix()
    except ValueError:
        return path.as_posix()


def collect_verification_rows(results_root: Path) -> List[Dict[str, Any]]:
    project_root = results_root.parent.parent
    rows: List[Dict[str, Any]] = []

    for folder in sorted(results_root.iterdir()):
        if not folder.is_dir() or folder.name == "identification":
            continue

        summary_path = folder / "summary.json"
        config_path = folder / "config.yaml"
        if not summary_path.exists() or not config_path.exists():
            continue

        summary = _load_json(summary_path)
        config = _load_yaml(config_path)
        modality = _infer_modality(folder.name, config)
        dataset_name = _infer_dataset_name(folder.name, config, modality)
        experiment_title = ""
        experiment = config.get("experiment", {})
        if isinstance(experiment, dict):
            title = experiment.get("name")
            if isinstance(title, str):
                experiment_title = title

        dataset_file = config.get("dataset", "")
        if not isinstance(dataset_file, str):
            dataset_file = ""

        evaluation_dir = folder / "evaluation"

        for matcher_name, metrics in summary.items():
            threshold_analysis_file = _find_threshold_analysis_file(evaluation_dir, matcher_name)
            auc_value = _extract_auc(threshold_analysis_file)
            eer_value = _extract_eer(evaluation_dir, matcher_name)

            row = {
                "experiment_folder": folder.name,
                "experiment_title": experiment_title,
                "base_matcher": _infer_base_matcher(matcher_name),
                "matcher": matcher_name,
                "modality": modality,
                "dataset": dataset_name,
                "dataset_file": dataset_file,
                "config_file": _format_relative_path(config_path, project_root),
                "summary_file": _format_relative_path(summary_path, project_root),
                "threshold_analysis_file": _format_relative_path(threshold_analysis_file, project_root)
                if threshold_analysis_file is not None
                else "",
                "eer": eer_value,
                "accuracy": metrics.get("accuracy"),
                "precision": metrics.get("precision"),
                "recall": metrics.get("recall"),
                "roc_auc": auc_value,
                "tar": metrics.get("tar"),
                "far": metrics.get("far"),
                "frr": metrics.get("frr"),
                "trr": metrics.get("trr"),
                "num_pairs": metrics.get("num_pairs"),
                "num_genuine": metrics.get("num_genuine"),
                "num_impostor": metrics.get("num_impostor"),
                "avg_inlier_ratio": metrics.get("avg_inlier_ratio"),
            }
            rows.append(row)

    return rows


def _cell_value(value: Any) -> Any:
    return value


def write_csv(rows: List[Dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: _cell_value(row.get(column, "")) for column in OUTPUT_COLUMNS})


def write_json(rows: List[Dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as handle:
        json.dump(rows, handle, indent=2, ensure_ascii=False)


def write_xlsx(rows: List[Dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Verification Results"

    header_font = Font(bold=True)
    worksheet.append(OUTPUT_COLUMNS)
    for cell in worksheet[1]:
        cell.font = header_font

    for row in rows:
        worksheet.append([_cell_value(row.get(column, "")) for column in OUTPUT_COLUMNS])

    for column_index, column_name in enumerate(OUTPUT_COLUMNS, start=1):
        if column_name in FLOAT_COLUMNS:
            for row_index in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_index, column=column_index).number_format = "0.###############"
        elif column_name in INTEGER_COLUMNS:
            for row_index in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_index, column=column_index).number_format = "0"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_index, column_name in enumerate(OUTPUT_COLUMNS, start=1):
        max_length = len(column_name)
        for cell in worksheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2, max_row=worksheet.max_row):
            for value_cell in cell:
                if value_cell.value is not None:
                    max_length = max(max_length, len(str(value_cell.value)))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 40)

    workbook.save(output_path)


def build_output_paths(output_base: Path) -> Dict[str, Path]:
    if output_base.suffix:
        stem = output_base.with_suffix("")
    else:
        stem = output_base

    return {
        "csv": stem.with_suffix(".csv"),
        "json": stem.with_suffix(".json"),
        "xlsx": stem.with_suffix(".xlsx"),
    }


def main(argv: Optional[Iterable[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Export verification experiment results to tabular files.")
    parser.add_argument(
        "--results-root",
        type=Path,
        default=_project_root() / "bioverify" / "results",
        help="Root directory with experiment result folders.",
    )
    parser.add_argument(
        "--output-base",
        type=Path,
        default=_project_root() / "bioverify" / "results" / "verification_results",
        help="Output file base path. CSV, JSON, and XLSX files will be written with this stem.",
    )
    args = parser.parse_args(list(argv) if argv is not None else None)

    rows = collect_verification_rows(args.results_root)
    if not rows:
        print(f"No verification experiment results found under {args.results_root}")
        return 1

    output_paths = build_output_paths(args.output_base)
    write_csv(rows, output_paths["csv"])
    write_json(rows, output_paths["json"])
    write_xlsx(rows, output_paths["xlsx"])

    print(f"Exported {len(rows)} verification rows")
    print(f"CSV:  {output_paths['csv']}")
    print(f"JSON: {output_paths['json']}")
    print(f"XLSX: {output_paths['xlsx']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())